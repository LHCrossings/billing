"""
aggregate.py - Read all MASTER FOR BILLING tabs from weekly log files,
combine into one dataset, validate col S (Month), correct any errors,
and standardize to the 15th of the billing month.
"""

import re
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl

LOGS_DIR = Path(__file__).parent / "logs"
SHEET_NAME = "MASTER FOR BILLING"

# Columns A through AC are the only relevant columns (indices 0-28)
LAST_COL_IDX = 28  # AC

# Column letter labels for indices 0-28
COL_LABELS = [
    "A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M",
    "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z",
    "AA", "AB", "AC",
]

# Key column indices
COL_B = 1   # Start Date
COL_N = 13  # Type (COM, PRG, BB, BNS, etc.)
COL_S = 18  # Month (True Month — we validate and standardize this)
COL_Y = 24  # Billing Type (Calendar / Broadcast)
COL_AA = 26 # Affidavit required (Y/N)
COL_AC = 28 # Market

MONTH_NAMES = [
    "", "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]

MONTH_ABBREVS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}


def broadcast_month_start(year: int, month: int) -> date:
    """Return the Monday of the week containing the 1st of the given month."""
    first = date(year, month, 1)
    return first - timedelta(days=first.weekday())  # weekday(): Mon=0, Sun=6


def expected_billing_month(start_date: date, billing_type: str) -> tuple[int, int]:
    """Return (year, month) for the correct billing month of a spot."""
    if billing_type.strip().lower() == "calendar":
        return (start_date.year, start_date.month)

    # Broadcast: find which broadcast month this date falls into
    for delta in range(-2, 3):
        month = start_date.month + delta
        year = start_date.year
        while month > 12:
            month -= 12
            year += 1
        while month < 1:
            month += 12
            year -= 1

        bm_start = broadcast_month_start(year, month)
        next_month = month % 12 + 1
        next_year = year + (1 if month == 12 else 0)
        bm_end = broadcast_month_start(next_year, next_month)

        if bm_start <= start_date < bm_end:
            return (year, month)

    return (start_date.year, start_date.month)  # fallback


def parse_month_col(value) -> tuple[int, int] | None:
    """Parse the Month column value to (year, month) regardless of format."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return (value.year, value.month)
    if isinstance(value, date):
        return (value.year, value.month)
    if isinstance(value, (int, float)):
        # Excel date serial: days since 1899-12-30
        d = date(1899, 12, 30) + timedelta(days=int(value))
        return (d.year, d.month)
    if isinstance(value, str):
        v = value.strip()
        # "Feb-26" style
        m = re.match(r"^([A-Za-z]{3})-(\d{2})$", v)
        if m:
            mon = MONTH_ABBREVS.get(m.group(1).lower())
            if mon:
                return (2000 + int(m.group(2)), mon)
        # "2/15/26" style
        m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2})$", v)
        if m:
            return (2000 + int(m.group(3)), int(m.group(1)))
        # "2/15/2026" style
        m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", v)
        if m:
            return (int(m.group(3)), int(m.group(1)))
    return None


def load_log(path: Path) -> tuple[list[dict], list[str]]:
    """
    Load MASTER FOR BILLING from one log file.
    Returns rows keyed by column letter (A-AC), plus a list of warnings.
    """
    warnings = []
    rows = []

    wb = openpyxl.load_workbook(path, read_only=True, keep_vba=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        warnings.append(f"Sheet '{SHEET_NAME}' not found in {path.name}")
        wb.close()
        return rows, warnings

    ws = wb[SHEET_NAME]
    first_row = True

    for row_vals in ws.iter_rows(values_only=True):
        # Only take columns A through AC
        row_vals = list(row_vals)[:LAST_COL_IDX + 1]
        # Pad if row is shorter than expected
        while len(row_vals) < LAST_COL_IDX + 1:
            row_vals.append(None)

        # Skip header row and blank rows
        if first_row:
            first_row = False
            continue
        if not any(v is not None for v in row_vals):
            continue

        record = {COL_LABELS[i]: row_vals[i] for i in range(LAST_COL_IDX + 1)}
        record["_source_file"] = path.name
        rows.append(record)

    wb.close()
    return rows, warnings


def validate_and_standardize(rows: list[dict]) -> tuple[list[dict], list[dict]]:
    """
    Validate col S (Month) against Start Date + Billing Type.
    Correct any mismatches and standardize all month values to the 15th
    of the correct billing month (e.g. date(2026, 2, 15)).

    Returns (corrected_rows, issues_found).
    """
    issues = []

    for row in rows:
        start_raw = row.get("B")  # Start Date
        billing_type = row.get("Y") or ""  # Billing Type
        month_raw = row.get("S")  # Month (raw, may be wrong/mixed format)

        # Determine the correct billing month
        if isinstance(start_raw, (date, datetime)) and billing_type.strip():
            start_date = start_raw.date() if isinstance(start_raw, datetime) else start_raw
            correct_year, correct_month = expected_billing_month(start_date, billing_type)
        else:
            # Can't compute — leave month as-is, just try to standardize format
            actual = parse_month_col(month_raw)
            if actual:
                correct_year, correct_month = actual
            else:
                # Nothing we can do — skip standardization
                continue

        # Check if col S matches what we computed
        actual = parse_month_col(month_raw)
        if actual is None:
            issues.append({
                "file": row["_source_file"],
                "bill_code": row.get("A"),
                "start_date": start_raw,
                "billing_type": billing_type,
                "month_raw": month_raw,
                "issue": "Month column could not be parsed — corrected from Start Date",
            })
        elif actual != (correct_year, correct_month):
            issues.append({
                "file": row["_source_file"],
                "bill_code": row.get("A"),
                "start_date": start_raw,
                "billing_type": billing_type,
                "month_raw": month_raw,
                "issue": (
                    f"Month said {MONTH_NAMES[actual[1]]} {actual[0]}, "
                    f"corrected to {MONTH_NAMES[correct_month]} {correct_year}"
                ),
            })

        # Standardize to the 15th of the correct billing month
        row["S"] = date(correct_year, correct_month, 15)

    return rows, issues


def filter_rows(rows: list[dict]) -> tuple[list[dict], list[dict], list[dict], list[dict]]:
    """
    Split rows into four buckets:
    - master: everything (no filtering)
    - billable: AA=Y and col A does not contain "DO NOT INVOICE"
    - dni_cleaned: DNI lines where bill code contains "broker" — go to CLEANED
    - dni_discard: all other DNI lines — master only, not carried forward
    """
    master = rows  # all rows, unchanged
    billable = []
    dni_cleaned = []
    dni_discard = []

    for row in rows:
        bill_code = str(row.get("A") or "").strip().upper()
        affidavit = str(row.get("AA") or "").strip().upper()

        if "DO NOT INVOICE" in bill_code:
            if "BROKER" in bill_code:
                dni_cleaned.append(row)
            else:
                dni_discard.append(row)
        elif affidavit == "Y":
            billable.append(row)
        # AA=N rows stay in master only

    return master, billable, dni_cleaned, dni_discard


def sort_rows(rows: list[dict]) -> list[dict]:
    """Sort billable rows: col A -> AB -> AC -> B (date) -> I (air time)."""
    def sort_key(row):
        a = str(row.get("A") or "")
        ab = str(row.get("AB") or "")
        ac = str(row.get("AC") or "")
        b = row.get("B")
        b = b.date() if isinstance(b, datetime) else (b or date.min)
        i = row.get("I")
        # Normalize air time for sorting: timedelta, string, or None
        if isinstance(i, timedelta):
            i_sort = i.total_seconds()
        elif isinstance(i, str):
            i_sort = i.lower()
        else:
            i_sort = 0
        return (a, ab, ac, b, i_sort)

    return sorted(rows, key=sort_key)


def has_contract(row: dict) -> bool:
    """Return True if col AB contains a valid Etere contract number."""
    ab = row.get("AB")
    if ab is None:
        return False
    ab_str = str(ab).strip()
    return ab_str not in ("", "0", "N")


def group_into_affidavits(rows: list[dict]) -> dict[str, list[dict]]:
    """
    Group sorted rows into affidavit buckets.
    - Has contract (AB valid)  -> key = contract number (AB)
    - No contract              -> key = bill code (A) + estimate (O)
    Returns an ordered dict of {affidavit_key: [rows]}.
    """
    groups: dict[str, list[dict]] = {}

    for row in rows:
        if has_contract(row):
            key = str(row["AB"]).strip()
        else:
            a = str(row.get("A") or "").strip()
            o = str(row.get("O") or "").strip()
            key = f"{a}|{o}" if o else a

        if key not in groups:
            groups[key] = []
        groups[key].append(row)

    return groups


def filter_by_month(rows: list[dict], year: int, month: int) -> list[dict]:
    """Keep only rows where col S falls in the given billing month."""
    return [
        row for row in rows
        if isinstance(row.get("S"), date)
        and row["S"].year == year
        and row["S"].month == month
    ]


def main():
    import sys
    # Billing month: pass as YYYY-MM argument, e.g. "2026-02". Defaults to current month.
    if len(sys.argv) > 1:
        year, month = map(int, sys.argv[1].split("-"))
    else:
        today = date.today()
        year, month = today.year, today.month
    print(f"Billing month: {MONTH_NAMES[month]} {year}\n")

    log_files = sorted(LOGS_DIR.glob("*.xlsm"))
    print(f"Found {len(log_files)} log files\n")

    all_rows = []
    all_warnings = []

    for path in log_files:
        rows, warnings = load_log(path)
        all_rows.extend(rows)
        all_warnings.extend(warnings)
        print(f"  {path.name}: {len(rows)} rows")

    print(f"\nTotal rows aggregated: {len(all_rows)}")

    if all_warnings:
        print(f"\n--- LOAD WARNINGS ({len(all_warnings)}) ---")
        for w in all_warnings:
            print(f"  {w}")

    all_rows, issues = validate_and_standardize(all_rows)

    if not issues:
        print("\nMonth validation: OK - no mismatches found")
    else:
        print(f"\n--- MONTH CORRECTIONS ({len(issues)}) ---")
        for issue in issues:
            print(
                f"  {issue['file']} | {issue['bill_code']} | "
                f"Start: {issue['start_date']} | {issue['billing_type']} | "
                f"Month col: {issue['month_raw']} -> {issue['issue']}"
            )

    all_rows_for_month = filter_by_month(all_rows, year, month)
    master, billable, dni_cleaned, dni_discard = filter_rows(all_rows_for_month)

    aa_n_count = len(master) - len(billable) - len(dni_cleaned) - len(dni_discard)
    print(f"\n--- ROW COUNTS AFTER FILTER ---")
    print(f"  Master (all rows):        {len(master)}")
    print(f"  Billable (AA=Y):          {len(billable)}")
    print(f"  DNI -> CLEANED (broker):  {len(dni_cleaned)}")
    print(f"  DNI -> discarded:         {len(dni_discard)}")
    print(f"  AA=N (master only):       {aa_n_count}")

    all_dni = dni_cleaned + dni_discard
    if all_dni:
        print(f"\n--- ALL DO NOT INVOICE LINES ---")
        for row in sorted(all_dni, key=lambda r: str(r.get("A") or "")):
            disposition = "-> CLEANED" if row in dni_cleaned else "-> discard"
            print(
                f"  [{disposition}] {row['_source_file']} | {row.get('A')} | "
                f"Start: {row.get('B')} | Gross: {row.get('P')} | "
                f"Market: {row.get('AC')}"
            )

    billable = sort_rows(billable)
    affidavits = group_into_affidavits(billable)

    # --- Worldlink ---
    from worldlink import load_all_worldlink
    wl_rows, wl_warnings = load_all_worldlink(billing_year=year, billing_month=month)

    if wl_warnings:
        print(f"\n--- WORLDLINK WARNINGS ({len(wl_warnings)}) ---")
        for w in wl_warnings:
            print(f"  {w}")

    wl_affidavits = group_into_affidavits(wl_rows)

    print(f"\n--- REGULAR AFFIDAVIT GROUPS ({len(affidavits)}) ---")
    print(f"  Total billable rows: {len(billable)}")
    print()
    for key, rows in affidavits.items():
        gross_total = sum(r.get("P") or 0 for r in rows)
        sample = rows[0]
        print(f"  [{key}] {sample.get('A')} | {sample.get('AC')} | {len(rows)} rows | ${gross_total:,.2f}")

    print(f"\n--- WORLDLINK AFFIDAVIT GROUPS ({len(wl_affidavits)}) ---")
    print(f"  Total Worldlink rows: {len(wl_rows)}")
    print()
    for key, rows in wl_affidavits.items():
        gross_total = sum(r.get("P") or 0 for r in rows)
        sample = rows[0]
        markets = sorted(set(r.get("Q") for r in rows))
        print(f"  [{key}] {sample.get('A')} | {', '.join(markets)} | {len(rows)} rows | ${gross_total:,.2f}")

    print(f"\n--- TOTALS ---")
    print(f"  Regular affidavits: {len(affidavits)}")
    print(f"  Worldlink affidavits: {len(wl_affidavits)}")
    print(f"  Combined: {len(affidavits) + len(wl_affidavits)}")


if __name__ == "__main__":
    main()
