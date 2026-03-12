"""
order_parser.py - Parse Sales Confirmation and Run Sheet from order Excel files.

Sales Confirmation column layout (0-indexed):
  Left side:  label at idx 1 (col B), value at idx 3 (col D)
  Right side: label at idx 8 (col I), value at idx 11 (col L) or idx 15 (col P) for financials
  City row:   city at idx 3, state at idx 5, zip at idx 6 (col B is blank)
  Monthly breakdown: after 'MONTHLY BREAKDOWN' marker — idx 1=month, idx 3=gross, idx 4=net
"""

import json
from datetime import date, datetime
from pathlib import Path

import openpyxl

from aggregate import COL_LABELS, LAST_COL_IDX, expected_billing_month, parse_month_col

MONTH_NAMES_TO_NUM = {
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
}


def _cell(row: list, idx: int, default=None):
    if idx >= len(row):
        return default
    return row[idx]


def _str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def parse_sales_confirmation(ws) -> dict:
    """
    Parse the Sales Confirmation sheet into a metadata dict.

    Returns keys: contract_number, client, advertiser, contact, address,
    city, state, zip, phone, fax, billing_type, market, estimate, notes,
    agency_discount, date_order_written, revision, station_rep, emails,
    total_gross, total_net, monthly_breakdown.

    monthly_breakdown is a list of {month_name, gross, net}.
    """
    result = {
        "contract_number": None,
        "client": None,
        "advertiser": None,
        "contact": None,
        "address": None,
        "city": None,
        "state": None,
        "zip": None,
        "phone": None,
        "fax": None,
        "billing_type": None,
        "market": None,
        "estimate": None,
        "notes": None,
        "agency_discount": 0.15,
        "date_order_written": None,
        "revision": None,
        "station_rep": None,
        "emails": [],
        "total_gross": None,
        "total_net": None,
        "monthly_breakdown": [],
    }

    all_rows = list(ws.iter_rows(values_only=True))
    in_header = True
    monthly_start_idx = None

    for row_idx, row in enumerate(all_rows):
        row = list(row)
        col_b = _str(_cell(row, 1))
        col_d = _cell(row, 3)
        col_f = _cell(row, 5)
        col_g = _cell(row, 6)
        col_i = _str(_cell(row, 8))
        col_k = _cell(row, 10)
        col_l = _cell(row, 11)

        if in_header:
            # Left-side labels (col B → value at col D)
            if col_b == "Client":
                result["client"] = _str(col_d)
            elif col_b == "Contact":
                result["contact"] = _str(col_d)
            elif col_b == "Address":
                result["address"] = _str(col_d)
            elif col_b == "" and col_f is not None and col_g is not None:
                # City/state/zip row (no label in col B)
                result["city"] = _str(col_d)
                result["state"] = _str(col_f)
                result["zip"] = _str(col_g)
            elif col_b == "Phone":
                result["phone"] = _str(col_d)
            elif col_b == "Fax":
                result["fax"] = _str(col_d)
            elif col_b == "Email":
                email = _str(col_d)
                if email:
                    result["emails"].append(email)
            elif col_b == "Station Representative":
                result["station_rep"] = _str(col_k)

            # Right-side labels (col I → value at col L)
            if col_i == "Advertiser":
                result["advertiser"] = _str(col_l)
            elif col_i == "Estimate":
                result["estimate"] = _str(col_l)
            elif col_i == "Billing Type":
                result["billing_type"] = _str(col_l)
            elif col_i == "Market":
                result["market"] = _str(col_l)
            elif col_i == "Date Order Written":
                v = col_l
                if isinstance(v, datetime):
                    result["date_order_written"] = v.strftime("%Y-%m-%d")
                elif isinstance(v, date):
                    result["date_order_written"] = v.isoformat()
                elif v:
                    result["date_order_written"] = _str(v)
            elif col_i == "Contract Number":
                try:
                    result["contract_number"] = int(col_l)
                except (TypeError, ValueError):
                    result["contract_number"] = col_l
            elif col_i == "Revision":
                try:
                    result["revision"] = int(col_l)
                except (TypeError, ValueError):
                    result["revision"] = col_l
            elif col_i == "Station Representative":
                result["station_rep"] = _str(col_k)

            if col_b == "Line Number":
                in_header = False
                continue

        # Financial totals appear after line items — detect in all rows post-header
        if not in_header:
            if col_i == "Gross Amount":
                # idx 15 = gross dollar total; idx 11 = spot count (skip)
                v = _cell(row, 15)
                if isinstance(v, (int, float)):
                    result["total_gross"] = float(v)
            elif col_i == "Agency Discount":
                # idx 11 = rate (0.15), idx 15 = dollar amount (negative)
                v = col_l
                if isinstance(v, float) and 0 < v < 1:
                    result["agency_discount"] = v
            elif col_i == "Net Amount of Contract":
                v = _cell(row, 15)
                if isinstance(v, (int, float)):
                    result["total_net"] = float(v)

            if col_b == "Additional Notes":
                notes = _str(col_d)
                if notes:
                    result["notes"] = notes

        # After header: look for MONTHLY BREAKDOWN marker
        for cell in row:
            if _str(cell) == "MONTHLY BREAKDOWN":
                monthly_start_idx = row_idx
                break

    # Parse monthly breakdown: skip 2 rows (blank + header), then read data rows
    if monthly_start_idx is not None:
        data_start = monthly_start_idx + 3  # marker + blank + header = +3 to reach first data row
        for row in all_rows[data_start:]:
            row = list(row)
            month_label = _str(_cell(row, 1))
            if not month_label or month_label.lower() == "total":
                break
            try:
                gross = float(_cell(row, 3) or 0)
            except (TypeError, ValueError):
                gross = 0.0
            try:
                net = float(_cell(row, 4) or 0)
            except (TypeError, ValueError):
                net = 0.0
            result["monthly_breakdown"].append({
                "month_name": month_label,
                "gross": gross,
                "net": net,
            })

    return result


def compute_monthly_from_runsheet(
    ws, billing_type: str, agency_discount: float
) -> list[dict]:
    """
    Compute monthly gross/net from Run Sheet col S (billing month) and col P (gross).
    Returns list of {year, month, gross, net}, sorted by (year, month).
    Skips header row and blank rows. Skips rows where P == 0.
    """
    # Key: (year, month, market, revenue_type)
    monthly: dict[tuple[int, int, str, str], float] = {}
    first_row = True

    for row_vals in ws.iter_rows(values_only=True):
        row_vals = list(row_vals)[: LAST_COL_IDX + 1]
        while len(row_vals) < LAST_COL_IDX + 1:
            row_vals.append(None)

        if first_row:
            first_row = False
            continue

        if not any(v is not None for v in row_vals):
            continue

        record = {COL_LABELS[i]: row_vals[i] for i in range(LAST_COL_IDX + 1)}

        gross_raw = record.get("P")
        if not isinstance(gross_raw, (int, float)) or gross_raw == 0:
            continue

        # Determine billing month from col S (preferred) or B + Y
        ym = parse_month_col(record.get("S"))
        if ym is None:
            b_raw = record.get("B")
            y_raw = record.get("Y") or billing_type
            if isinstance(b_raw, (date, datetime)):
                b_date = b_raw.date() if isinstance(b_raw, datetime) else b_raw
                ym = expected_billing_month(b_date, y_raw)
            else:
                continue

        year, month = ym
        market = str(record.get("AC") or "").strip() or "Unknown"
        revenue_type = str(record.get("X") or "").strip()
        key = (year, month, market, revenue_type)
        monthly[key] = monthly.get(key, 0.0) + float(gross_raw)

    result = []
    for (year, month, market, revenue_type), gross in sorted(monthly.items()):
        net = gross * (1 - agency_discount)
        result.append({
            "year": year, "month": month, "market": market,
            "revenue_type": revenue_type, "gross": gross, "net": net,
        })
    return result


def parse_order_file(path: Path, metadata_only: bool = False) -> dict | None:
    """
    Parse one order Excel file.
    Returns a dict ready for orders_db.upsert_order / upsert_monthly, or None if
    the file is not an order file (missing required sheets or contract number).

    The returned dict contains a '_monthly' key with the per-month breakout list.
    Callers must pop '_monthly' before passing to upsert_order.

    If metadata_only=True, the Run Sheet is skipped and '_monthly' is always empty.
    Use this for Worldlink orders where monthly revenue comes from Etere CSVs instead.
    """
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return None

    if "Sales Confirmation" not in wb.sheetnames:
        wb.close()
        return None

    if not metadata_only and "Run Sheet" not in wb.sheetnames:
        wb.close()
        return None

    meta = parse_sales_confirmation(wb["Sales Confirmation"])
    if not isinstance(meta.get("contract_number"), int):
        wb.close()
        return None

    billing_type = meta["billing_type"] or "Broadcast"
    agency_discount = meta["agency_discount"] or 0.15

    if metadata_only:
        monthly = []
    else:
        monthly = compute_monthly_from_runsheet(wb["Run Sheet"], billing_type, agency_discount)
    wb.close()

    cn = meta["contract_number"]
    for row in monthly:
        row["contract_number"] = cn

    return {
        "contract_number": cn,
        "file_path": str(path),
        "client": meta["client"],
        "advertiser": meta["advertiser"],
        "contact": meta["contact"],
        "address": meta["address"],
        "city": meta["city"],
        "state": meta["state"],
        "zip": meta["zip"],
        "phone": meta["phone"],
        "fax": meta["fax"],
        "billing_type": meta["billing_type"],
        "market": meta["market"],
        "estimate": meta["estimate"],
        "notes": meta["notes"],
        "agency_discount": agency_discount,
        "date_order_written": meta["date_order_written"],
        "revision": meta["revision"],
        "station_rep": meta["station_rep"],
        "emails": json.dumps(meta["emails"]),
        "total_gross": meta["total_gross"],
        "total_net": meta["total_net"],
        "last_updated": datetime.now().isoformat(),
        "_monthly": monthly,
    }
